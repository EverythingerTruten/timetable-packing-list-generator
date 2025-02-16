import openpyxl
from openpyxl.styles import PatternFill
from tkinter import filedialog
import tkinter as tk

root = tk.Tk()
root.withdraw()

#Setup preferences

path = filedialog.askopenfilename(title='Select spreadsheet file',
    filetypes=[
        ("XLSX files", ".xlsx")
    ]
)

preference_path = filedialog.askopenfilename(title='Select preference file',
    filetypes=[
        ("Text files", ".txt")
    ]
)

file = open(preference_path, "r")
lines = []
for x in file:
    lines.append(x)
file.close()

ignored_classes_line = lines[9]
ignored_classes_stripped = ignored_classes_line.rstrip()
ignored_classes = ignored_classes_stripped.split(',')

lis_line = lines[12]
left_in_school = lis_line.rstrip()

def cell_to_coord(cell_ref):
    col_str = ''.join(c for c in cell_ref if c.isalpha())
    row = int(''.join(c for c in cell_ref if c.isdigit()))    
    col = 0
    for i, letter in enumerate(reversed(col_str)):
        col += (ord(letter.upper()) - ord('A') + 1) * (26 ** i)    
    return (row, col)

if input('Does your timetable include class numbers? (y/n): ') == 'y':
    class_num_toggle = True
else:
    class_num_toggle = False

wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
while True:
    subject_area_codes = input('Input the area with subject names (For example A1/C4 means the area between the cell A1 and C4): ')
    corner_cords = []
    for i in subject_area_codes.split('/'):
        corner_cords.append(cell_to_coord(i))

    column_count = corner_cords[1][1] - corner_cords[0][1]
    if column_count == 4:
        break
    else:
        print("The input area doesn't have 5 columns. Please try again")
        continue

row_count = corner_cords[1][0] - corner_cords[0][0] + 1

def day_plan_to_list(row_count, first_row, column):
    list = []
    for i in range(row_count):
        cell = sheet_obj.cell(row = first_row + i, column = column)
        list.append(cell.value)
    return list

def list_cleanup(input_list, remove_after_space):
    filtered_list = [x for x in input_list if x]

    if remove_after_space:
        filtered_list = [str(x).split(' ')[0] for x in filtered_list]
    
    seen = set()
    result = []
    for item in filtered_list:
        if item not in seen:
            seen.add(item)
            result.append(item)
            
    return result

def has_pe(day_list):
    return left_in_school in day_list

#Generate the packing list
first_row = corner_cords[0][0]
for i in range(5):
    tomorrow_column = corner_cords[0][1] + i
    clean_tomorrow_list = list_cleanup(day_plan_to_list(row_count, first_row, tomorrow_column), class_num_toggle)

    # Get day after tomorrow for PE logic
    day_after_tomorrow_column = tomorrow_column + 1
    if day_after_tomorrow_column <= corner_cords[1][1]:
        clean_day_after_tomorrow = list_cleanup(day_plan_to_list(row_count, first_row, day_after_tomorrow_column), class_num_toggle)
    else:
        clean_day_after_tomorrow = []

    if tomorrow_column == corner_cords[0][1]:
        today = corner_cords[1][1]
    else:
        today = tomorrow_column - 1
    
    clean_today = list_cleanup(day_plan_to_list(row_count, first_row, today), class_num_toggle)

    # Generate take-out list
    take_out_list = []
    
    # Handle regular subjects
    for a in clean_today:
        if a in clean_tomorrow_list:
            continue
        elif a in ignored_classes:
            continue
        elif a == left_in_school:
            continue
        else:
            take_out_list.append(a)
    
    # Handle PE for take-out list - check tomorrow vs day after tomorrow
    if has_pe(clean_tomorrow_list) and not has_pe(clean_day_after_tomorrow):
        take_out_list.append(left_in_school)
    
    # Generate pack list
    pack_list = []
    
    # Handle PE for pack list first
    if not has_pe(clean_today) and has_pe(clean_tomorrow_list):
        pack_list.append(left_in_school)
    
    # Handle regular subjects
    for a in clean_tomorrow_list:
        if a in clean_today:
            continue
        elif a in ignored_classes:
            continue
        elif a == left_in_school:
            continue
        else:
            pack_list.append(a)

    print(f"Take out: {take_out_list}")
    print(f"Pack: {pack_list}")
    print("---")

    #Save lists to .xlsx file

    red_color = 'e06666'
    green_color = '93c47d'

    for a in range(len(take_out_list)):
        writing_cell = sheet_obj.cell(first_row + row_count + a + 1, corner_cords[0][1] + i)
        writing_cell.value = take_out_list[a] + ' -'
        writing_cell.fill = PatternFill(start_color=red_color, end_color=red_color, fill_type="solid")

    for a in range(len(pack_list)):
        writing_cell = sheet_obj.cell(first_row + row_count + 1 + len(take_out_list) + a, corner_cords[0][1] + i)
        writing_cell.value = pack_list[a] + ' +'
        writing_cell.fill = PatternFill(start_color=green_color, end_color=green_color, fill_type="solid")

wb_obj.save(path)